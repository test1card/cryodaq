"""Экспорт данных из SQLite в Excel (.xlsx) через openpyxl."""

from __future__ import annotations

import logging
import math
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from cryodaq.storage.archive_reader import ArchiveReader
from cryodaq.storage.sentinel import decode
from cryodaq.storage.sqlite_writer import _parse_timestamp

logger = logging.getLogger(__name__)

_XLSX_MAX_ROWS = 1_048_576


def _ts_sort_key(raw: object) -> float:
    """Comparable sort key for mixed REAL/legacy-ISO timestamp keys.

    A data dir may mix REAL(float) epoch timestamps and legacy ISO-string
    timestamps; ``sorted`` on the raw keys raises TypeError. Normalize each
    to an epoch float first.
    """
    try:
        return _parse_timestamp(raw).timestamp()
    except (ValueError, TypeError, OSError):
        return float("inf")


class XLSXExporter:
    """Экспортирует данные из SQLite daily-файлов в Excel .xlsx.

    Формат:
    - Лист 1 "Данные": время | канал1 | канал2 | ...
    - Лист 2 "Информация": метаданные эксперимента

    Пример использования::

        exporter = XLSXExporter(data_dir=Path("data"))
        count = exporter.export(
            output_path=Path("export/readings.xlsx"),
            start=datetime(2026, 3, 14, tzinfo=timezone.utc),
            end=datetime(2026, 3, 15, tzinfo=timezone.utc),
        )
    """

    def __init__(self, data_dir: Path, archive_dir: Path | None = None) -> None:
        """Инициализировать экспортёр.

        Параметры
        ----------
        data_dir:
            Директория с daily-файлами SQLite (data_YYYY-MM-DD.db).
        archive_dir:
            Директория холодного хранилища (Parquet + index.json). None →
            ``data_dir / "archive"``. Дни, вытесненные ротацией в Parquet,
            читаются оттуда, иначе экспорт слепнет на вытесненных днях.
        """
        self._data_dir = data_dir
        self._archive_dir = archive_dir if archive_dir is not None else data_dir / "archive"

    def export(
        self,
        output_path: Path,
        *,
        start: datetime | None = None,
        end: datetime | None = None,
        channels: list[str] | None = None,
        experiment_info: dict[str, Any] | None = None,
    ) -> int:
        """Экспортировать показания в .xlsx.

        Параметры
        ----------
        output_path:
            Путь для создаваемого XLSX-файла.
        start:
            Начало временного диапазона (включительно).  None → без ограничения.
        end:
            Конец временного диапазона (исключительно).  None → без ограничения.
        channels:
            Фильтр по именам каналов.  None → все каналы.
        experiment_info:
            Дополнительные метаданные для листа "Информация".

        Возвращает
        ----------
        int:  Количество экспортированных временных меток (строк данных).
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            logger.error("openpyxl не установлен: pip install openpyxl")
            return 0

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()

        # ------------------------------------------------------------------
        # Лист 1: Данные
        # ------------------------------------------------------------------
        ws_data = wb.active
        ws_data.title = "Данные"

        # Read across hot SQLite + cold Parquet: rotated days live only in the
        # archive, so a plain SQLite scan would drop them silently.
        rows = ArchiveReader(self._data_dir, self._archive_dir).query_rows(start, end, channels, None)
        all_rows: list[dict[str, Any]] = [
            {"timestamp": raw_ts, "channel": channel, "value": value, "status": status}
            for raw_ts, _instrument_id, channel, value, _unit, status in rows
        ]

        if not all_rows:
            wb.save(str(output_path))
            logger.info("XLSX экспорт: %s (0 записей, нет данных)", output_path)
            return 0

        # Уникальные каналы в алфавитном порядке
        unique_channels = sorted({r["channel"] for r in all_rows})

        # Заголовок
        bold = Font(bold=True)
        headers = ["Время"] + unique_channels
        for col, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            cell.font = bold

        # Сводка по времени: ts_str → {channel: value}
        # NaN-доктрина: decode at the read boundary — a sentinel / error / legacy
        # ±inf row surfaces as NaN and is left as an empty cell below, never as a
        # non-physical number in the operator's spreadsheet.
        by_time: dict[str, dict[str, float]] = defaultdict(dict)
        for r in all_rows:
            by_time[r["timestamp"]][r["channel"]] = decode(r["value"], r["status"])

        row_num = 2
        for ts_str in sorted(by_time.keys(), key=_ts_sort_key):
            if row_num >= _XLSX_MAX_ROWS:
                logger.warning(
                    "XLSX row limit reached (%d). Truncating export — some data omitted.",
                    _XLSX_MAX_ROWS,
                )
                break
            # Разобрать REAL или ISO-строку в datetime для Excel
            try:
                dt = _parse_timestamp(ts_str)
                # Excel не понимает offset-aware datetime — приводим к UTC naive
                dt = dt.astimezone(UTC).replace(tzinfo=None)
            except (ValueError, TypeError, OSError):
                dt = ts_str  # fallback: оставить как строку

            ts_cell = ws_data.cell(row=row_num, column=1, value=dt)
            if isinstance(dt, datetime):
                ts_cell.number_format = "YYYY-MM-DD HH:MM:SS"

            channel_values = by_time[ts_str]
            for col, ch in enumerate(unique_channels, 2):
                v = channel_values.get(ch)
                if v is not None and math.isfinite(v):
                    # Write the full float value (no pre-rounding): vacuum
                    # pressures span 1e-3..1e-9 mbar and were collapsed to 0.000
                    # by round(v, 3) + "0.000" format. "General" lets Excel pick
                    # a representation that preserves both small and wide-range
                    # magnitudes.
                    cell = ws_data.cell(row=row_num, column=col, value=v)
                    cell.number_format = "General"

            row_num += 1

        data_row_count = row_num - 2

        # ------------------------------------------------------------------
        # Лист 2: Информация
        # ------------------------------------------------------------------
        ws_info = wb.create_sheet("Информация")
        info_rows: list[tuple[str, Any]] = [
            ("Система", "CryoDAQ"),
            ("Дата экспорта", datetime.now().isoformat()),
            ("Записей", data_row_count),
            ("Каналов", len(unique_channels)),
        ]
        if start is not None:
            info_rows.append(("Начало диапазона", start.isoformat()))
        if end is not None:
            info_rows.append(("Конец диапазона", end.isoformat()))
        if experiment_info:
            for k, v in experiment_info.items():
                info_rows.append((str(k), str(v)))

        for row_idx, (key, val) in enumerate(info_rows, 1):
            ws_info.cell(row=row_idx, column=1, value=key).font = bold
            ws_info.cell(row=row_idx, column=2, value=val)

        # ------------------------------------------------------------------
        # Авторазмер столбцов
        # ------------------------------------------------------------------
        for ws in (ws_data, ws_info):
            for col_cells in ws.columns:
                max_len = max(
                    (len(str(cell.value or "")) for cell in col_cells),
                    default=10,
                )
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)

        wb.save(str(output_path))
        logger.info(
            "XLSX экспорт: %s (%d записей, %d каналов)",
            output_path,
            data_row_count,
            len(unique_channels),
        )
        return data_row_count
