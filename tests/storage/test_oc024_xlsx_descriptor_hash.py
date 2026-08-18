from __future__ import annotations

import sqlite3
from contextlib import closing
from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from cryodaq.channels.descriptors import (
    ChannelCatalog,
    ChannelDescriptorV1,
    ChannelQuantity,
    ChannelRole,
    ChannelSafetyClass,
)
from cryodaq.storage.channel_descriptors import initialize_descriptor_storage, install_catalog
from cryodaq.storage.sqlite_writer import SCHEMA_READINGS
from cryodaq.storage.xlsx_export import XLSXExporter


def _descriptor(channel: str) -> ChannelDescriptorV1:
    return ChannelDescriptorV1(
        schema_version=1,
        channel_id=channel,
        instrument_id="i",
        source_key="input.temperature",
        quantity=ChannelQuantity.TEMPERATURE,
        unit="K",
        role=ChannelRole.PRIMARY_MEASUREMENT,
        safety_class=ChannelSafetyClass.OBSERVATIONAL,
        display_group="Cryostat",
        display_name="Термопара",
        visible_by_default=True,
        display_order=1,
        descriptor_revision=1,
    )


def _database(path: Path, *, channel: str, descriptor: ChannelDescriptorV1 | None, legacy: bool) -> None:
    """Write one day's DB. ``descriptor`` None + legacy builds the pre-catalog
    shape (no descriptor_hash column); ``descriptor`` None + not legacy builds a
    dangling-hash shape (column present, no catalog row) that exporters must refuse."""
    with closing(sqlite3.connect(path)) as conn:
        conn.execute(SCHEMA_READINGS)
        values = (
            datetime.fromisoformat(path.stem.removeprefix("data_")).replace(tzinfo=UTC).timestamp(),
            "i",
            channel,
            1.0,
            "K",
            "ok",
        )
        if descriptor is not None:
            initialize_descriptor_storage(conn)
            install_catalog(conn, ChannelCatalog([descriptor]))
            conn.execute(
                "INSERT INTO readings "
                "(timestamp, instrument_id, channel, value, unit, status, descriptor_hash) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (*values, descriptor.descriptor_hash),
            )
        elif not legacy:
            conn.execute("ALTER TABLE readings ADD COLUMN descriptor_hash TEXT")
            conn.execute(
                "INSERT INTO readings "
                "(timestamp, instrument_id, channel, value, unit, status, descriptor_hash) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (*values, "sha256:" + "a" * 64),
            )
        else:
            conn.execute(
                "INSERT INTO readings "
                "(timestamp, instrument_id, channel, value, unit, status) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                values,
            )
        # closing() closes the handle; it does NOT commit.
        conn.commit()


def test_xlsx_exports_descriptor_hash_and_leaves_old_archive_empty(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    descriptor = _descriptor("CH_NEW")
    _database(data_dir / "data_2026-03-14.db", channel="CH_NEW", descriptor=descriptor, legacy=False)
    _database(data_dir / "data_2026-03-15.db", channel="CH_OLD", descriptor=None, legacy=True)

    output_path = tmp_path / "identity.xlsx"
    XLSXExporter(data_dir).export(output_path)

    ws = openpyxl.load_workbook(output_path)["Данные"]
    assert [ws.cell(row=1, column=column).value for column in range(1, 4)] == [
        "Время",
        "i",
        "i",
    ]
    assert [ws.cell(row=2, column=column).value for column in range(2, 4)] == [
        "CH_NEW",
        "CH_OLD",
    ]
    assert ws.cell(row=3, column=2).value == descriptor.descriptor_hash
    assert ws.cell(row=3, column=3).value is None, "old archive descriptor_hash cell must be empty"


def test_xlsx_exports_fail_closed_on_dangling_descriptor_hash(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    _database(data_dir / "data_2026-03-14.db", channel="CH_NEW", descriptor=None, legacy=False)

    output_path = tmp_path / "must-not-exist.xlsx"
    from cryodaq.storage.archive_reader import ArchiveUnavailableError

    try:
        XLSXExporter(data_dir).export(output_path)
    except ArchiveUnavailableError as caught:
        assert "descriptor_hash_missing" in str(caught)
    else:
        raise AssertionError("dangling-hash XLSX export must fail closed")
    assert not output_path.exists()
