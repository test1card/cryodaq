"""Tests for XLSXExporter — SQLite → XLSX export with sheet structure and filters."""

from __future__ import annotations

import json
import math
import sqlite3
from contextlib import closing
from datetime import UTC, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import pytest

from cryodaq.drivers.base import ChannelStatus, Reading
from cryodaq.storage.sentinel import SENTINEL
from cryodaq.storage.sqlite_writer import SQLiteWriter
from cryodaq.storage.xlsx_export import XLSXExporter

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _reading(
    channel: str = "CH1",
    value: float = 4.5,
    unit: str = "K",
    *,
    ts: datetime,
    instrument_id: str = "ls218s",
    status: ChannelStatus = ChannelStatus.OK,
) -> Reading:
    return Reading(
        timestamp=ts,
        instrument_id=instrument_id,
        channel=channel,
        value=value,
        unit=unit,
        status=status,
    )


def _populate_db(data_dir: Path, readings: list[Reading]) -> None:
    """Write readings directly into the data_dir via SQLiteWriter._write_batch()."""
    writer = SQLiteWriter(data_dir)
    writer._write_batch(readings)
    if writer._conn is not None:
        writer._conn.close()  # close, don't just drop the ref — Windows locks open DB files
    writer._conn = None


# ---------------------------------------------------------------------------
# 1. export() creates an XLSX file and returns a positive row count
# ---------------------------------------------------------------------------


async def test_xlsx_creates_file(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    readings = [_reading("CH1", 4.5, ts=ts) for _ in range(5)]
    _populate_db(data_dir, readings)

    output_path = tmp_path / "out.xlsx"
    exporter = XLSXExporter(data_dir)
    count = exporter.export(output_path)

    assert output_path.exists(), "XLSX file was not created"
    assert count > 0, f"Expected count > 0, got {count}"


# ---------------------------------------------------------------------------
# 2. Workbook has exactly two sheets: "Данные" and "Информация"
# ---------------------------------------------------------------------------


async def test_xlsx_has_two_sheets(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    _populate_db(data_dir, [_reading(ts=ts)])

    output_path = tmp_path / "out.xlsx"
    XLSXExporter(data_dir).export(output_path)

    wb = openpyxl.load_workbook(output_path)
    assert wb.sheetnames == ["Данные", "Информация"], f"Unexpected sheet names: {wb.sheetnames}"


# ---------------------------------------------------------------------------
# 3. "Данные" sheet contains correct headers and known data values
# ---------------------------------------------------------------------------


async def test_xlsx_data_values_correct(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)

    readings = [
        _reading("CH1", 4.5, "K", ts=ts),
        _reading("CH2", 77.0, "K", ts=ts),
    ]
    _populate_db(data_dir, readings)

    output_path = tmp_path / "out.xlsx"
    XLSXExporter(data_dir).export(output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Данные"]

    # Channel identities occupy three header rows.
    channels = [cell.value for cell in ws[2]]
    assert "CH1" in channels, f"CH1 missing from channel row: {channels}"
    assert "CH2" in channels, f"CH2 missing from channel row: {channels}"

    ch1_col = channels.index("CH1") + 1
    ch2_col = channels.index("CH2") + 1

    # The first data row follows the three identity header rows.
    ch1_val = ws.cell(row=4, column=ch1_col).value
    ch2_val = ws.cell(row=4, column=ch2_col).value

    assert ch1_val is not None and abs(ch1_val - 4.5) < 0.01, f"CH1 value: {ch1_val}"
    assert ch2_val is not None and abs(ch2_val - 77.0) < 0.01, f"CH2 value: {ch2_val}"


# ---------------------------------------------------------------------------
# 4. Empty DB → export returns 0 and file is created with no data rows
# ---------------------------------------------------------------------------


async def test_xlsx_empty_db_returns_zero(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    # Initialise the DB schema without inserting any readings
    writer = SQLiteWriter(data_dir)
    writer._ensure_connection(datetime(2026, 3, 14).date())
    if writer._conn is not None:
        writer._conn.close()  # close, don't just drop the ref — Windows locks open DB files
    writer._conn = None

    output_path = tmp_path / "out.xlsx"
    exporter = XLSXExporter(data_dir)
    count = exporter.export(output_path)

    assert count == 0, f"Expected 0 rows from empty DB, got {count}"
    assert output_path.exists(), "XLSX should be created even when DB is empty"

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Данные"]
    # Row 1 is the header; no data rows expected
    assert ws.max_row <= 1, f"Expected no data rows, found max_row={ws.max_row}"


# ---------------------------------------------------------------------------
# 5. Channel filter — only requested channels appear in the output
# ---------------------------------------------------------------------------


async def test_xlsx_channel_filter(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)

    readings = [
        _reading("CH1", 4.5, "K", ts=ts),
        _reading("CH2", 77.0, "K", ts=ts),
        _reading("CH3", 300.0, "K", ts=ts),
    ]
    _populate_db(data_dir, readings)

    output_path = tmp_path / "out.xlsx"
    count = XLSXExporter(data_dir).export(output_path, channels=["CH1"])

    assert count >= 1, f"Expected at least 1 row for CH1 filter, got {count}"

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Данные"]

    # The channel identity row should only contain CH1 (not CH2, CH3).
    channels = [cell.value for cell in ws[2]]
    assert "CH1" in channels, f"CH1 missing from channel row: {channels}"
    assert "CH2" not in channels, f"CH2 should not be in filtered output: {channels}"
    assert "CH3" not in channels, f"CH3 should not be in filtered output: {channels}"


# ---------------------------------------------------------------------------
# 6. _XLSX_MAX_ROWS constant is defined and has the correct value
# ---------------------------------------------------------------------------


async def test_xlsx_max_rows_constant(tmp_path: Path) -> None:
    """_XLSX_MAX_ROWS is the Excel hard limit; monkeypatch a small cap and
    confirm the exporter actually truncates output rather than writing past it.
    """
    import unittest.mock as _mock

    import cryodaq.storage.xlsx_export as xlsx_mod

    # Canonical value correct
    assert xlsx_mod._XLSX_MAX_ROWS == 1_048_576, f"Expected _XLSX_MAX_ROWS == 1_048_576, got {xlsx_mod._XLSX_MAX_ROWS}"

    # Runtime truncation: write 10 readings at distinct timestamps (10 data rows)
    data_dir = tmp_path / "data"
    ts_base = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    readings = [_reading("CH1", float(i), ts=ts_base.replace(second=i)) for i in range(10)]
    _populate_db(data_dir, readings)

    output_path = tmp_path / "out_capped.xlsx"
    # cap=5: data starts at row 4 after three identity header rows; the loop
    # breaks at row 5, so exactly one data row can be written.
    cap = 5
    with _mock.patch.object(xlsx_mod, "_XLSX_MAX_ROWS", cap):
        exporter = XLSXExporter(data_dir)
        count = exporter.export(output_path)

    assert output_path.exists(), "XLSX not created under small cap"
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Данные"]
    data_rows = ws.max_row - 3  # subtract the three identity header rows
    assert data_rows < 10, f"Exporter wrote {data_rows} data rows despite cap={cap}; truncation broken"
    assert count == data_rows, f"Return value {count} must equal actual data rows written {data_rows}"


# ---------------------------------------------------------------------------
# 7. ME-9 / D-C8 — small vacuum-range values must not be truncated to 0.000
# ---------------------------------------------------------------------------


async def test_xlsx_preserves_small_vacuum_values(tmp_path: Path) -> None:
    """A 1e-7 mbar reading must round-trip as a non-zero cell value.

    Regression: exporter pre-rounded to 3 decimals with number_format "0.000",
    collapsing vacuum pressures (1e-3..1e-9 mbar) to 0.000.
    """
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    _populate_db(data_dir, [_reading("VAC", 1e-7, "mbar", ts=ts)])

    output_path = tmp_path / "vac.xlsx"
    XLSXExporter(data_dir).export(output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Данные"]
    channels = [cell.value for cell in ws[2]]
    col = channels.index("VAC") + 1
    val = ws.cell(row=4, column=col).value

    assert val is not None, "vacuum cell is empty"
    assert val != 0.0, f"vacuum value truncated to zero: {val}"
    assert abs(val - 1e-7) < 1e-12, f"vacuum value not preserved: {val}"


def test_xlsx_preserves_same_channel_readings_with_distinct_descriptor_hashes(tmp_path: Path) -> None:
    """A timestamp/channel collision must retain each descriptor identity."""
    from cryodaq.channels.descriptors import (
        ChannelCatalog,
        ChannelDescriptorV1,
        ChannelQuantity,
        ChannelRole,
        ChannelSafetyClass,
    )
    from cryodaq.storage.channel_descriptors import install_catalog

    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    _populate_db(
        data_dir,
        [
            _reading("CH1", 4.5, ts=ts),
            _reading("CH1", 7.5, ts=ts),
        ],
    )

    def _descriptor(revision: int) -> ChannelDescriptorV1:
        return ChannelDescriptorV1(
            schema_version=1,
            channel_id="CH1",
            instrument_id="ls218s",
            source_key="input.ch1.temperature",
            quantity=ChannelQuantity.TEMPERATURE,
            unit="K",
            role=ChannelRole.PRIMARY_MEASUREMENT,
            safety_class=ChannelSafetyClass.OBSERVATIONAL,
            display_group="Cryostat",
            display_name="Термопара",
            visible_by_default=True,
            display_order=1,
            descriptor_revision=revision,
        )

    first = _descriptor(revision=1)
    second = _descriptor(revision=2)
    db_path = data_dir / f"data_{ts.date().isoformat()}.db"
    with closing(sqlite3.connect(str(db_path))) as conn:
        install_catalog(conn, ChannelCatalog([first]))
        install_catalog(conn, ChannelCatalog([second], historical=[first, second]))
        row_ids = [row[0] for row in conn.execute("SELECT id FROM readings ORDER BY id")]
        assert len(row_ids) == 2
        conn.execute("UPDATE readings SET descriptor_hash=? WHERE id=?", (first.descriptor_hash, row_ids[0]))
        conn.execute("UPDATE readings SET descriptor_hash=? WHERE id=?", (second.descriptor_hash, row_ids[1]))
        conn.commit()

    output_path = tmp_path / "descriptor-collision.xlsx"
    assert XLSXExporter(data_dir).export(output_path) == 1

    ws = openpyxl.load_workbook(output_path).active
    values_by_hash = {
        ws.cell(row=3, column=column).value: ws.cell(row=4, column=column).value
        for column in range(2, ws.max_column + 1)
        if ws.cell(row=1, column=column).value == "ls218s" and ws.cell(row=2, column=column).value == "CH1"
    }

    assert values_by_hash == {first.descriptor_hash: 4.5, second.descriptor_hash: 7.5}

    ws_info = openpyxl.load_workbook(output_path)["Информация"]
    rows = {
        ws_info.cell(row=row, column=1).value: ws_info.cell(row=row, column=2).value
        for row in range(1, ws_info.max_row + 1)
    }
    assert rows.get("Идентичностей") == 2, (
        "two descriptor-qualified identities for one channel must be reported as identities"
    )
    assert "Каналов" not in rows, "the info sheet must not present descriptor identities as channels"


# ---------------------------------------------------------------------------
# 8. D-C18 — mixed TEXT (legacy ISO) + REAL timestamps must not raise
# ---------------------------------------------------------------------------


async def test_xlsx_mixed_timestamp_types(tmp_path: Path) -> None:
    """Sorting mixed str/float timestamp keys must not raise TypeError."""
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    _populate_db(data_dir, [_reading("CH1", 1.0, "K", ts=ts)])  # REAL ts row

    db_path = data_dir / f"data_{ts.date().isoformat()}.db"
    conn = sqlite3.connect(str(db_path))
    conn.execute(
        "INSERT INTO readings (timestamp, instrument_id, channel, value, unit, status) VALUES (?, ?, ?, ?, ?, ?)",
        ("2026-03-14T13:00:00+00:00", "ls218s", "CH1", 2.0, "K", "ok"),
    )
    conn.commit()
    conn.close()

    output_path = tmp_path / "mixed.xlsx"
    count = XLSXExporter(data_dir).export(output_path)  # must not raise TypeError
    assert count == 2, f"expected 2 exported timestamps, got {count}"


# ---------------------------------------------------------------------------
# 9. D-C9 / ME-10 — early local-hours range must select correct UTC day file
# ---------------------------------------------------------------------------


async def test_xlsx_selects_utc_day_for_early_local_hours(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts = datetime(2026, 3, 13, 22, 30, 0, tzinfo=UTC)  # data_2026-03-13.db
    _populate_db(data_dir, [_reading("CH1", 1.0, "K", ts=ts)])

    msk = timezone(timedelta(hours=3))
    start = datetime(2026, 3, 14, 0, 0, 0, tzinfo=msk)  # 2026-03-13 21:00 UTC
    end = datetime(2026, 3, 14, 6, 0, 0, tzinfo=msk)  # 2026-03-14 03:00 UTC

    output_path = tmp_path / "early.xlsx"
    count = XLSXExporter(data_dir).export(output_path, start=start, end=end)
    assert count == 1, f"early-hours UTC day file dropped: got {count}"


# ---------------------------------------------------------------------------
# NaN-доктрина: a stored sentinel/error reading must never surface as a number
# ---------------------------------------------------------------------------


def test_xlsx_masks_sentinel_row(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    ts_good = datetime(2026, 3, 14, 12, 0, 0, tzinfo=UTC)
    ts_bad = datetime(2026, 3, 14, 12, 0, 30, tzinfo=UTC)
    _populate_db(
        data_dir,
        [
            _reading("CH1", 4.5, ts=ts_good, status=ChannelStatus.OK),
            _reading("CH1", float("nan"), ts=ts_bad, status=ChannelStatus.SENSOR_ERROR),
        ],
    )

    output_path = tmp_path / "masked.xlsx"
    XLSXExporter(data_dir).export(output_path)

    ws = openpyxl.load_workbook(str(output_path))["Данные"]
    values = [c.value for row in ws.iter_rows(min_row=2) for c in row[1:]]
    assert SENTINEL not in values, "sentinel leaked into XLSX"
    assert not any(isinstance(v, float) and not math.isfinite(v) for v in values), "non-finite number leaked into XLSX"
    assert 4.5 in values, "usable reading must still be exported"


def test_xlsx_export_rejects_missing_indexed_source_before_writing(tmp_path: Path) -> None:
    data_dir = tmp_path / "data"
    archive_dir = data_dir / "archive"
    archive_dir.mkdir(parents=True)
    (archive_dir / "index.json").write_text(
        json.dumps(
            {
                "files": [
                    {
                        "original_name": "data_2026-03-14.db",
                        "archive_path": "year=2026/month=03/data_2026-03-14.db.parquet",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    output_path = tmp_path / "must-not-exist.xlsx"

    with pytest.raises(RuntimeError) as caught:
        XLSXExporter(data_dir).export(output_path)

    assert type(caught.value).__name__ == "ArchiveUnavailableError"
    assert not output_path.exists()
